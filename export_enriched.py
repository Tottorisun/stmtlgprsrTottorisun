# -*- coding: utf-8 -*-
"""
Экспорт лидов ВМЕСТЕ с правовой формой (ИП/ООО/...) в один .xlsx для владельца.

Штатный main.py --export-only пишет только таблицу leads; правовая форма лежит
в отдельной таблице legal_form_enrichment (так задумано — обогащение аддитивно
и не трогает пайплайн). Этот скрипт соединяет их для человека, который работает
в Excel, а не в SQL:

  - лист «ИП — горячие»   : только ip с уверенностью high/medium (первыми в работу)
  - лист «Все лиды»       : все 1905 с колонками формы/уверенности/метода
  - лист «На проверку»    : строки с пометкой возможного межгородского дубля
                            или «затекания» города — то, что просили глазами
  - лист «Сводка»         : разбивка по регионам ИП/ООО/гос/прочее/неизвестно

Ничего не пишет в базу, только читает. Запуск:
    python export_enriched.py
"""
import sqlite3
import sys
import io
from collections import defaultdict
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DB = ROOT / "data" / "leads.sqlite3"
OUT = ROOT / "out" / "стоматологии_ИП_обогащение.xlsx"

FORM_RU = {"ip": "ИП", "ooo": "ООО", "gov": "гос/муниципальная",
           "other": "прочее (АО/АНО/…)", "unknown": "неизвестно"}
CONF_RU = {"high": "высокая", "medium": "средняя", "low": "низкая", "none": "—"}

COLS = [
    ("org_name", "Организация", 30), ("region", "Регион", 18), ("city", "Город", 16),
    ("address", "Адрес", 34), ("phone", "Телефон(ы)", 20),
    ("legal_form", "Правовая форма", 18), ("confidence", "Уверенность", 12),
    ("matched_name", "Найдено в реестре", 26), ("matched_inn", "ИНН", 14),
    ("matched_ogrn", "ОГРН/ОГРНИП", 16), ("method", "Метод", 26),
    ("notes", "Примечания", 40),
]
HDR_FILL = PatternFill("solid", fgColor="2F5597")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=9.5)
WRAP = Alignment(wrap_text=True, vertical="top")


def rows():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    q = ("SELECT l.org_name, l.region, l.city, l.address, l.phone, l.notes, "
         "       e.legal_form, e.confidence, e.matched_name, e.matched_inn, "
         "       e.matched_ogrn, e.method "
         "FROM leads l LEFT JOIN legal_form_enrichment e ON e.dedupe_key = l.dedupe_key "
         "ORDER BY l.region, l.city, l.org_name")
    out = []
    for r in conn.execute(q):
        d = dict(r)
        d["legal_form"] = FORM_RU.get(d["legal_form"], d["legal_form"] or "")
        d["confidence"] = CONF_RU.get(d["confidence"], d["confidence"] or "")
        out.append(d)
    conn.close()
    return out


def sheet(wb, title, data):
    ws = wb.create_sheet(title[:31])
    ws.append([c[1] for c in COLS])
    for d in data:
        ws.append([d.get(c[0]) for c in COLS])
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font, c.alignment = BODY, WRAP
    for i, (_, _, w) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(ws.max_row, 1)}"
    return ws


def summary(wb, data):
    ws = wb.create_sheet("Сводка", 0)
    ws.append(["Регион", "Всего", "ИП", "ООО", "гос", "прочее", "неизвестно"])
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
    per = defaultdict(lambda: defaultdict(int))
    for d in data:
        per[d["region"]][d["legal_form"]] += 1
    tot = defaultdict(int)
    order = sorted(per, key=lambda r: -sum(per[r].values()))
    for region in order:
        p = per[region]
        n = sum(p.values())
        ws.append([region, n, p[FORM_RU["ip"]], p[FORM_RU["ooo"]], p[FORM_RU["gov"]],
                   p[FORM_RU["other"]], p[FORM_RU["unknown"]]])
        tot["n"] += n
        for k in ("ip", "ooo", "gov", "other", "unknown"):
            tot[k] += p[FORM_RU[k]]
    ws.append(["ИТОГО", tot["n"], tot["ip"], tot["ooo"], tot["gov"], tot["other"], tot["unknown"]])
    for c in ws[ws.max_row]:
        c.font = Font(name="Arial", bold=True)
    for i, w in enumerate([22, 8, 8, 8, 8, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    data = rows()
    hot = [d for d in data if d["legal_form"] == FORM_RU["ip"]
           and d["confidence"] in (CONF_RU["high"], CONF_RU["medium"])]
    review = [d for d in data if d.get("notes") and
              ("межгородск" in d["notes"] or "артефакт районного" in d["notes"]
               or "город карточки" in d["notes"])]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary(wb, data)
    sheet(wb, "ИП — горячие", hot)
    sheet(wb, "Все лиды", data)
    sheet(wb, "На проверку", review)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Готово: {OUT}")
    print(f"  всего лидов: {len(data)}; ИП высокой/средней уверенности: {len(hot)}; "
          f"на ручную проверку (дубли/город): {len(review)}")


if __name__ == "__main__":
    main()
