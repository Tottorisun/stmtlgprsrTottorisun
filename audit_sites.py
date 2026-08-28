# -*- coding: utf-8 -*-
"""
CLI аудита сайтов клиник на соответствие обязательному составу информации по
закону (разворот 28.08.2026). Работает по базе, собранной в режиме has-site
(python main.py --mode has-site ...): берёт лиды, у которых есть сайт, и по
каждому сайту прогоняет src/site_audit.py.

Примеры:
    # проаудировать все сайты в базе has-site
    python audit_sites.py

    # PoC на одном регионе, ограничить 60 клиниками
    python audit_sites.py --region "Тюменская область" --limit 60

    # другая база (напр. изолированная под параллельный сбор)
    python audit_sites.py --db-path data/parallel/tyumen_hassite.sqlite3

    # только пересобрать отчёт/‌.xlsx из уже накопленного site_audit, без сети
    python audit_sites.py --report-only

ЧЕК-ПОЙНТ (жёсткое правило владельца, как в src/pipeline.py): результат каждой
клиники пишется в БД и коммитится СРАЗУ, до перехода к следующей. Прерывание
процесса теряет максимум одну клинику, а не весь прогон. Дозапуск (--resume, по
умолчанию) пропускает уже успешно проверенные сайты.

ЧЕСТНОСТЬ: каждый «отсутствует» — ГИПОТЕЗА («вероятно отсутствует — проверить
вручную»), не факт (см. docstring src/site_audit.py). Отчёт это проговаривает.
"""
import argparse
import sys
import io
import time
import random
from pathlib import Path

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import requests

from src import db
from src.site_audit import (
    audit_site, ensure_audit_table, save_audit, already_audited, CHECKS, UA,
)

DEFAULT_DB = ROOT / "data" / "leads_with_site.sqlite3"
DEFAULT_XLSX = ROOT / "out" / "аудит_сайтов.xlsx"


def fetch_site_leads(conn, region=None):
    q = ("SELECT dedupe_key, org_name, region, city, address, phone, email, website, "
         "source_url FROM leads WHERE has_website = 1 AND COALESCE(website, '') <> ''")
    params = []
    if region:
        q += " AND region = ?"
        params.append(region)
    q += " ORDER BY region, city, org_name"
    cur = conn.execute(q, params)
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def run_audit(conn, leads, resume=True, pause=(0.4, 0.9), log=print):
    # Пауза МЕЖДУ клиниками маленькая намеренно: соседние клиники — это РАЗНЫЕ
    # домены, так что нагрузку на один сервер она не создаёт. Вежливость, которая
    # реально важна (не долбить ОДИН сайт), обеспечивается паузой между 2-3
    # запросами ВНУТРИ одного сайта (pause=(0.8,1.6) в audit_site) и жёстким
    # потолком «максимум 3 запроса на домен». Большая пауза между разными
    # доменами только замедляла бы прогон без выигрыша в вежливости.
    ensure_audit_table(conn)
    done = already_audited(conn) if resume else set()
    session = requests.Session()
    session.headers.update({"User-Agent": UA, "Accept-Language": "ru-RU,ru;q=0.9",
                            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8"})
    robots_cache = {}
    n_total = len(leads)
    n_done = n_ok = 0
    for i, lead in enumerate(leads, 1):
        if resume and lead["dedupe_key"] in done:
            continue
        url = lead.get("website") or ""
        log(f"[{i}/{n_total}] {lead['org_name'][:40]:40s} | {url[:50]}")
        result = audit_site(url, session=session, robots_cache=robots_cache,
                            log=log, pause=(0.8, 1.6))
        save_audit(conn, lead, result, db.now_iso())
        conn.commit()  # <-- ЧЕК-ПОЙНТ: клиника сохранена, что бы ни случилось дальше
        n_done += 1
        if result["fetched_ok"]:
            n_ok += 1
            thin = " [тонкий/JS-контент — проверить вручную]" if result.get("thin_content") else ""
            log(f"      score {result['score']}/{result['max_score']} | "
                f"вероятно отсутствует: {', '.join(result['missing']) or '—'}{thin}")
        else:
            log(f"      сайт не открылся: {result['error']}")
        time.sleep(random.uniform(*pause))
    session.close()
    return n_done, n_ok


def report(conn, log=print):
    """Честная сводка распределения из таблицы site_audit."""
    cur = conn.execute("SELECT COUNT(*), SUM(fetched_ok), SUM(thin_content) FROM site_audit")
    total, ok, thin = cur.fetchone()
    total = total or 0
    ok = ok or 0
    thin = thin or 0
    log("\n" + "=" * 66)
    log("СВОДКА АУДИТА САЙТОВ (по таблице site_audit)")
    log("=" * 66)
    log(f"Всего записей аудита:        {total}")
    log(f"Сайт открылся (reachable):   {ok}"
        + (f"  ({round(100*ok/total)}%)" if total else ""))
    log(f"Не открылся:                 {total - ok}")
    log(f"Тонкий/JS-контент (ненадёжно): {thin} — по ним «отсутствует» особенно "
        f"подозрительно, нужна ручная проверка")
    if ok:
        log(f"\nРаспределение ВЕРОЯТНО ОТСУТСТВУЮЩИХ элементов (из {ok} открывшихся сайтов):")
        for key, spec in CHECKS.items():
            c = _missing_count(conn, key)
            pct = round(100 * c / ok) if ok else 0
            caveat = "  [!] низкая точность детекта" if spec.get("caveat") else ""
            log(f"  {pct:3d}%  ({c:4d})  {spec['label']}{caveat}")
        # средний «score»
        avg = conn.execute("SELECT AVG(score) FROM site_audit WHERE fetched_ok = 1").fetchone()[0]
        log(f"\nСредний балл соответствия: {avg:.1f} / {len(CHECKS)} "
            f"(по эвристике; выше = больше обязательных элементов НАЙДЕНО)")
    log("\nВАЖНО: каждый «вероятно отсутствует» — ГИПОТЕЗА, не факт. Куки-баннер и")
    log("версия для слабовидящих чаще всего грузятся JavaScript'ом и по статике")
    log("детектируются хуже всего — их «отсутствие» обязательно проверять вручную")
    log("на живом сайте, прежде чем писать клинике.")


def _missing_count(conn, key):
    """Точный счётчик клиник, у которых элемент key в списке missing."""
    cur = conn.execute("SELECT missing FROM site_audit WHERE fetched_ok = 1")
    n = 0
    for (missing,) in cur.fetchall():
        keys = [k.strip() for k in (missing or "").split(";") if k.strip()]
        if key in keys:
            n += 1
    return n


def export_xlsx(conn, out_path, log=print):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    check_keys = list(CHECKS.keys())
    headers = (["Организация", "Регион", "Город", "Сайт", "Открылся", "Балл",
                "Тонкий/JS"] + [CHECKS[k]["label"] for k in check_keys]
               + ["Вероятно отсутствует (для письма)", "Ошибка"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аудит сайтов"
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="2F5597")
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    import json as _json
    cur = conn.execute(
        "SELECT org_name, region, city, url, final_url, fetched_ok, score, max_score, "
        "thin_content, missing_labels, checks_json, error FROM site_audit "
        "ORDER BY region, city, org_name")
    for (org, region, city, url, final_url, ok, score, maxs, thin, miss_labels,
         checks_json, error) in cur.fetchall():
        checks = {}
        try:
            checks = _json.loads(checks_json or "{}")
        except Exception:
            pass
        row = [org, region, city, final_url or url, "да" if ok else "нет",
               f"{score}/{maxs}" if ok else "", "да" if thin else ""]
        for k in check_keys:
            v = checks.get(k) or {}
            if not ok:
                row.append("")
            elif v.get("present"):
                row.append("есть")
            else:
                row.append("нет?")  # гипотеза
        row.append(miss_labels or "")
        row.append(error or "")
        ws.append(row)

    widths = [30, 18, 14, 34, 9, 7, 9] + [16] * len(check_keys) + [50, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(name="Arial", size=9.5)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row,1)}"

    # лист-легенда с правовой привязкой и оговоркой о честности
    ws2 = wb.create_sheet("Легенда и оговорка")
    ws2.append(["Элемент", "Правовая привязка", "Оговорка по детекту"])
    for c in ws2[1]:
        c.fill, c.font = hdr_fill, hdr_font
    for k in check_keys:
        ws2.append([CHECKS[k]["label"], CHECKS[k]["law"],
                    CHECKS[k].get("caveat") or "—"])
    ws2.append([])
    ws2.append(["ВАЖНО", "Каждый «нет?» — ГИПОТЕЗА (вероятно отсутствует), не факт. "
                "Проверять вручную на живом сайте перед письмом клинике.", ""])
    for i, w in enumerate([40, 46, 60], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in ws2.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log(f"\n.xlsx отчёт аудита: {out_path}")
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--db-path", default=str(DEFAULT_DB),
                    help=f"база, собранная в режиме has-site (по умолчанию {DEFAULT_DB})")
    ap.add_argument("--region", default=None, help="проверять только этот регион (точное имя)")
    ap.add_argument("--limit", type=int, default=None, help="проверить не более N клиник (для PoC)")
    ap.add_argument("--no-resume", action="store_true",
                    help="перепроверить все, включая уже проверенные (по умолчанию — дозапуск)")
    ap.add_argument("--report-only", action="store_true",
                    help="не ходить в сеть — только сводка/‌.xlsx из накопленного site_audit")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="путь для .xlsx-отчёта")
    args = ap.parse_args()

    db_path = Path(args.db_path)
    if not db_path.exists():
        sys.exit(f"Нет базы has-site: {db_path}. Сначала соберите её: "
                 f"python main.py --mode has-site --region <id>")
    conn = db.connect(str(db_path))
    ensure_audit_table(conn)

    if args.report_only:
        report(conn)
        export_xlsx(conn, Path(args.xlsx))
        conn.close()
        return

    leads = fetch_site_leads(conn, region=args.region)
    if args.limit:
        leads = leads[:args.limit]
    print(f"К аудиту: {len(leads)} клиник с сайтом"
          + (f" в регионе {args.region!r}" if args.region else "")
          + (f" (лимит {args.limit})" if args.limit else ""))
    if not leads:
        print("Нет клиник с сайтом для аудита. Соберите базу: "
              "python main.py --mode has-site --region <id>")
        conn.close()
        return

    n_done, n_ok = run_audit(conn, leads, resume=not args.no_resume)
    print(f"\nПроверено в этом прогоне: {n_done} | открылись: {n_ok}")
    report(conn)
    export_xlsx(conn, Path(args.xlsx))
    conn.close()


if __name__ == "__main__":
    main()
