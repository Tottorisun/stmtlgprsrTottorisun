# -*- coding: utf-8 -*-
"""
.xlsx-экспорт в том же стиле, что и в RUSIMEX/leadgen (build_contacts_book.py):
закреплённая шапка, автофильтр, читаемые ширины колонок, отдельный лист-сводка.
"""
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import FIELDS

HDR_RU = {
    "dedupe_key": "Ключ дедупликации",
    "org_name": "Организация",
    "category": "Категория",
    "region": "Регион",
    "city": "Город",
    "address": "Адрес",
    "phone": "Телефон(ы)",
    "email": "Email",
    "has_website": "Есть сайт",
    "source": "Источник(и)",
    "source_url": "Ссылка(и) на карточку",
    "scraped_at": "Собрано",
    "notes": "Примечания",
}
WIDTHS = {
    "dedupe_key": 10, "org_name": 30, "category": 22, "region": 18, "city": 14,
    "address": 34, "phone": 20, "email": 22, "has_website": 10, "source": 16,
    "source_url": 40, "scraped_at": 18, "notes": 24,
}
HDR_FILL = PatternFill("solid", fgColor="2F5597")
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=9.5)
WRAP = Alignment(wrap_text=True, vertical="top")


def _write_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.append([HDR_RU[f] for f in FIELDS])
    for r in rows:
        row = [r.get(f) for f in FIELDS]
        row[FIELDS.index("has_website")] = "нет"  # база по построению — только без сайта
        ws.append(row)
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY_FONT
            c.alignment = WRAP
    for i, f in enumerate(FIELDS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(f, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FIELDS))}{max(ws.max_row, 1)}"
    return ws


def _write_summary(wb, rows_by_region):
    ws = wb.create_sheet("Сводка", 0)
    ws.append(["Регион", "Лидов (без сайта)", "Есть телефон", "Есть email", "Только email (без тел.)"])
    for c in ws[1]:
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center")
    total = Counter()
    for region, rows in rows_by_region.items():
        with_phone = sum(1 for r in rows if r.get("phone"))
        with_email = sum(1 for r in rows if r.get("email"))
        email_only = sum(1 for r in rows if r.get("email") and not r.get("phone"))
        ws.append([region, len(rows), with_phone, with_email, email_only])
        total["n"] += len(rows)
        total["phone"] += with_phone
        total["email"] += with_email
        total["email_only"] += email_only
    ws.append(["ИТОГО", total["n"], total["phone"], total["email"], total["email_only"]])
    for c in ws[ws.max_row]:
        c.font = Font(name="Arial", bold=True)
    for i, w in enumerate([22, 18, 14, 14, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def export_xlsx(rows_by_region, out_path):
    """rows_by_region: dict {имя_региона: [lead-dict, ...]}. Пишет один .xlsx:
    лист 'Сводка' + один лист на регион."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_summary(wb, rows_by_region)
    for region, rows in rows_by_region.items():
        _write_sheet(wb, region[:31], rows)  # Excel: имя листа <= 31 символ
    wb.save(out_path)
    return out_path
